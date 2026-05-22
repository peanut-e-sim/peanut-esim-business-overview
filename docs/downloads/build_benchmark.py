from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HOLAFLY_STD = ("Unlimited 7d", "Unlimited", "7 days", 27.30)
HOLAFLY_HEAVY = ("Unlimited 30d", "Unlimited", "30 days", 74.90)

DATA = {
    "Australia": {"Airalo":{"std":("Movida 7d","5GB","7 days",10.00),"heavy":("Movida 30d","10GB","30 days",18.00)},"Holafly":{"std":HOLAFLY_STD,"heavy":HOLAFLY_HEAVY},"Nomad":{"std":("Aussie 30d","5GB","30 days",12.00),"heavy":("Aussie 30d","10GB","30 days",18.00)},"Saily":{"std":("Standard 30d","5GB","30 days",10.99),"heavy":("Best Choice 30d","10GB","30 days",17.99)}},
    "China": {"Airalo":{"std":("Chinacom 7d","5GB","7 days",14.50),"heavy":("Chinacom 30d","10GB","30 days",26.50)},"Holafly":{"std":HOLAFLY_STD,"heavy":HOLAFLY_HEAVY},"Nomad":{"std":("China 30d","5GB","30 days",13.00),"heavy":("China 30d sale","10GB","30 days",16.00)},"Saily":{"std":("Standard 30d","5GB","30 days",15.99),"heavy":("Standard 30d","10GB","30 days",26.99)}},
    "Hong Kong": {"Airalo":{"std":("Webbing 7d","5GB","7 days",11.00),"heavy":("Webbing 30d","10GB","30 days",18.00)},"Holafly":{"std":HOLAFLY_STD,"heavy":HOLAFLY_HEAVY},"Nomad":{"std":("HK 30d","5GB","30 days",11.00),"heavy":("HK 30d","10GB","30 days",17.00)},"Saily":{"std":("Standard 30d","5GB","30 days",11.99),"heavy":("Best Choice 30d","10GB","30 days",19.99)}},
    "Japan": {"Airalo":{"std":("Moshi Moshi 7d","5GB","7 days",10.00),"heavy":("Moshi Moshi 30d","10GB","30 days",18.00)},"Holafly":{"std":HOLAFLY_STD,"heavy":HOLAFLY_HEAVY},"Nomad":{"std":("Japan 30d","5GB","30 days",10.00),"heavy":("Japan 30d sale","10GB","30 days",17.00)},"Saily":{"std":("Standard 30d","5GB","30 days",10.99),"heavy":("Best Choice 30d","10GB","30 days",17.99)}},
    "New Zealand": {"Airalo":{"std":("Kiwi 7d","5GB","7 days",13.00),"heavy":("Kiwi 30d","10GB","30 days",24.00)},"Holafly":{"std":HOLAFLY_STD,"heavy":HOLAFLY_HEAVY},"Nomad":{"std":("NZ 30d","5GB","30 days",14.00),"heavy":("NZ 30d","10GB","30 days",21.00)},"Saily":{"std":("Standard 30d","5GB","30 days",13.99),"heavy":("Best Choice 30d","10GB","30 days",23.99)}},
    "South Korea": {"Airalo":{"std":("Jum 7d","5GB","7 days",10.00),"heavy":("Jum 30d","10GB","30 days",19.00)},"Holafly":{"std":HOLAFLY_STD,"heavy":HOLAFLY_HEAVY},"Nomad":{"std":("Korea 30d","5GB","30 days",11.00),"heavy":("Korea 30d","10GB","30 days",18.00)},"Saily":{"std":("Standard 30d","5GB","30 days",10.99),"heavy":("Standard 30d","10GB","30 days",18.99)}},
    "Taiwan": {"Airalo":{"std":("Twn 7d","5GB","7 days",11.00),"heavy":("Twn 30d","10GB","30 days",19.00)},"Holafly":{"std":HOLAFLY_STD,"heavy":HOLAFLY_HEAVY},"Nomad":{"std":("Taiwan 30d","5GB","30 days",11.00),"heavy":("Taiwan 30d","10GB","30 days",18.00)},"Saily":{"std":("Standard 30d","5GB","30 days",11.99),"heavy":("Best Choice 30d","10GB","30 days",19.99)}},
    "South Africa": {"Airalo":{"std":("Sasa 7d","5GB","7 days",11.50),"heavy":("Sasa 30d","10GB","30 days",22.00)},"Holafly":{"std":HOLAFLY_STD,"heavy":HOLAFLY_HEAVY},"Nomad":{"std":("SA 30d","5GB","30 days",13.00),"heavy":("SA 30d","10GB","30 days",22.00)},"Saily":{"std":("Standard 30d","5GB","30 days",12.99),"heavy":("Standard 30d","10GB","30 days",22.99)}},
    "Canada": {"Airalo":{"std":("Cancom 7d","5GB","7 days",14.00),"heavy":("Cancom 30d","10GB","30 days",23.00)},"Holafly":{"std":HOLAFLY_STD,"heavy":HOLAFLY_HEAVY},"Nomad":{"std":("Canada 30d","5GB","30 days",14.00),"heavy":("Canada 30d","10GB","30 days",23.00)},"Saily":{"std":("Standard 30d","5GB","30 days",17.99),"heavy":("Best Choice 30d","10GB","30 days",30.99)}},
    "Mexico": {"Airalo":{"std":("Mexcom 7d","5GB","7 days",12.00),"heavy":("Mexcom 30d","10GB","30 days",18.50)},"Holafly":{"std":HOLAFLY_STD,"heavy":HOLAFLY_HEAVY},"Nomad":{"std":("Mexico 30d","5GB","30 days",12.00),"heavy":("Mexico 30d","10GB","30 days",19.00)},"Saily":{"std":("Standard 30d","5GB","30 days",16.99),"heavy":("Best Choice 30d","10GB","30 days",24.99)}},
    "USA": {"Airalo":{"std":("Change 7d","5GB","7 days",13.00),"heavy":("Change 30d","10GB","30 days",23.00)},"Holafly":{"std":HOLAFLY_STD,"heavy":HOLAFLY_HEAVY},"Nomad":{"std":("USA 30d","5GB","30 days",13.00),"heavy":("USA 30d","10GB","30 days",21.00)},"Saily":{"std":("Standard 30d","5GB","30 days",13.99),"heavy":("Best Choice 30d","10GB","30 days",22.99)}},
    "Argentina": {"Airalo":{"std":("Mate 7d","5GB","7 days",16.00),"heavy":("Mate 30d","10GB","30 days",30.00)},"Holafly":{"std":HOLAFLY_STD,"heavy":HOLAFLY_HEAVY},"Nomad":{"std":("Argentina 30d","5GB","30 days",16.00),"heavy":("Argentina 30d","10GB","30 days",28.00)},"Saily":{"std":("Standard 30d","5GB","30 days",16.99),"heavy":("Best Choice 30d","10GB","30 days",29.99)}},
    "Brazil": {"Airalo":{"std":("Brava 7d","5GB","7 days",13.00),"heavy":("Brava 30d","10GB","30 days",25.00)},"Holafly":{"std":HOLAFLY_STD,"heavy":HOLAFLY_HEAVY},"Nomad":{"std":("Brazil 30d","5GB","30 days",14.00),"heavy":("Brazil 30d","10GB","30 days",23.00)},"Saily":{"std":("Standard 30d","5GB","30 days",13.99),"heavy":("Best Choice 30d","10GB","30 days",24.99)}},
    "Colombia": {"Airalo":{"std":("Cafe 7d","5GB","7 days",15.00),"heavy":("Cafe 30d","10GB","30 days",28.00)},"Holafly":{"std":HOLAFLY_STD,"heavy":HOLAFLY_HEAVY},"Nomad":{"std":("Colombia 30d","5GB","30 days",14.00),"heavy":("Colombia 30d","10GB","30 days",25.00)},"Saily":{"std":("Standard 30d","5GB","30 days",16.99),"heavy":("Best Choice 30d","10GB","30 days",29.99)}},
    "UK": {"Airalo":{"std":("Uki 7d","5GB","7 days",12.00),"heavy":("Uki 30d","10GB","30 days",19.00)},"Holafly":{"std":HOLAFLY_STD,"heavy":HOLAFLY_HEAVY},"Nomad":{"std":("UK 30d","5GB","30 days",12.00),"heavy":("UK 30d","10GB","30 days",19.00)},"Saily":{"std":("Standard 30d","5GB","30 days",12.99),"heavy":("Best Choice 30d","10GB","30 days",18.99)}},
    "France": {"Airalo":{"std":("Bouquet 7d","5GB","7 days",9.50),"heavy":("Bouquet 30d","10GB","30 days",16.00)},"Holafly":{"std":HOLAFLY_STD,"heavy":HOLAFLY_HEAVY},"Nomad":{"std":("France 30d","5GB","30 days",11.00),"heavy":("France 30d","10GB","30 days",18.00)},"Saily":{"std":("Standard 30d","5GB","30 days",11.99),"heavy":("Best Choice 30d","10GB","30 days",19.99)}},
    "Italy": {"Airalo":{"std":("Iditaly 7d","5GB","7 days",12.00),"heavy":("Iditaly 30d","10GB","30 days",20.50)},"Holafly":{"std":HOLAFLY_STD,"heavy":HOLAFLY_HEAVY},"Nomad":{"std":("Italy 30d","5GB","30 days",11.00),"heavy":("Italy 30d","10GB","30 days",18.00)},"Saily":{"std":("Standard 30d","5GB","30 days",12.99),"heavy":("Standard 30d","10GB","30 days",20.99)}},
    "Spain": {"Airalo":{"std":("Espana 7d","5GB","7 days",9.00),"heavy":("Espana 30d","10GB","30 days",15.50)},"Holafly":{"std":HOLAFLY_STD,"heavy":HOLAFLY_HEAVY},"Nomad":{"std":("Spain 30d","5GB","30 days",11.00),"heavy":("Spain 30d","10GB","30 days",17.00)},"Saily":{"std":("Standard 30d","5GB","30 days",9.99),"heavy":("Standard 30d","10GB","30 days",15.99)}},
    "Thailand": {"Airalo":{"std":("True 7d","5GB","7 days",7.00),"heavy":("True 30d","10GB","30 days",11.00)},"Holafly":{"std":HOLAFLY_STD,"heavy":HOLAFLY_HEAVY},"Nomad":{"std":("Thailand 30d","5GB","30 days",8.00),"heavy":("Thailand 30d","10GB","30 days",12.00)},"Saily":{"std":("Standard 30d","5GB","30 days",7.99),"heavy":("Standard 30d","10GB","30 days",10.99)}},
    "Singapore": {"Airalo":{"std":("Connect 7d","5GB","7 days",9.00),"heavy":("Connect 30d","10GB","30 days",16.00)},"Holafly":{"std":HOLAFLY_STD,"heavy":HOLAFLY_HEAVY},"Nomad":{"std":("Singapore 30d","5GB","30 days",10.00),"heavy":("Singapore 30d","10GB","30 days",16.00)},"Saily":{"std":("Standard 30d","5GB","30 days",9.99),"heavy":("Best Choice 30d","10GB","30 days",15.99)}},
    "India": {"Airalo":{"std":("Jeevan 7d","5GB","7 days",13.00),"heavy":("Jeevan 30d","10GB","30 days",24.00)},"Holafly":{"std":HOLAFLY_STD,"heavy":HOLAFLY_HEAVY},"Nomad":{"std":("India 30d","5GB","30 days",13.00),"heavy":("India 30d","10GB","30 days",22.00)},"Saily":{"std":("Standard 30d","5GB","30 days",13.99),"heavy":("Standard 30d","10GB","30 days",23.99)}},
    "Pakistan": {"Airalo":{"std":("Zong 7d","5GB","7 days",10.00),"heavy":("Zong 30d","10GB","30 days",18.00)},"Holafly":{"std":HOLAFLY_STD,"heavy":HOLAFLY_HEAVY},"Nomad":{"std":("Pakistan 30d","5GB","30 days",11.00),"heavy":("Pakistan 30d","10GB","30 days",18.00)},"Saily":{"std":("Standard 30d","5GB","30 days",10.99),"heavy":("Best Choice 30d","10GB","30 days",17.99)}},
}

DESTINATIONS = list(DATA.keys())
COMPETITORS = ["Airalo","Holafly","Nomad","Saily"]
TIERS = {"Australia":"A","China":"A","Hong Kong":"A","Japan":"A","New Zealand":"A","South Korea":"A","Taiwan":"A","South Africa":"A","Canada":"A","Mexico":"A","USA":"A","Argentina":"A","Brazil":"A","Colombia":"A","UK":"B","France":"B","Italy":"B","Spain":"B","Thailand":"B","Singapore":"B","India":"C","Pakistan":"C"}
SLUGS = {"Australia":"australia","China":"china","Hong Kong":"hong-kong","Japan":"japan","New Zealand":"new-zealand","South Korea":"south-korea","Taiwan":"taiwan","South Africa":"south-africa","Canada":"canada","Mexico":"mexico","USA":"united-states","Argentina":"argentina","Brazil":"brazil","Colombia":"colombia","UK":"united-kingdom","France":"france","Italy":"italy","Spain":"spain","Thailand":"thailand","Singapore":"singapore","India":"india","Pakistan":"pakistan"}
HOLAFLY_SLUGS = dict(SLUGS); HOLAFLY_SLUGS["USA"]="usa"; HOLAFLY_SLUGS["UK"]="united-kingdom"

def url_for(comp, dest):
    if comp=="Airalo": return f"https://www.airalo.com/{SLUGS[dest]}-esim"
    if comp=="Holafly": return f"https://esim.holafly.com/esim-{HOLAFLY_SLUGS[dest]}/"
    if comp=="Nomad": return f"https://www.nomadesim.com/{SLUGS[dest]}-eSIM"
    if comp=="Saily": return f"https://saily.com/esim-{SLUGS[dest]}/"

wb = Workbook()
HEADER_FILL = PatternFill("solid", start_color="1F2937")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Arial", size=10)
DEST_FONT = Font(name="Arial", bold=True, size=10)
THIN = Side(border_style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GREEN_FILL = PatternFill("solid", start_color="D1FAE5")
RED_FILL = PatternFill("solid", start_color="FEE2E2")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

def style_header(ws, row, last_col):
    for c in range(1, last_col+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = CENTER; cell.border = BORDER

def build_pricing(ws, key):
    headers = ["Destination","Airalo Plan","Airalo Price (USD)","Holafly Plan","Holafly Price (USD)","Nomad Plan","Nomad Price (USD)","Saily Plan","Saily Price (USD)","Notes"]
    ws.append(headers); style_header(ws, 1, 10)
    for i, dest in enumerate(DESTINATIONS, start=2):
        row = [dest]
        notes = []
        for comp in COMPETITORS:
            entry = DATA[dest][comp][key]
            plan_name, data, duration, price = entry
            row.extend([f"{plan_name} ({data}/{duration})", price])
            if comp=="Holafly":
                notes.append("Holafly: unlimited, day-based only")
            elif comp in ("Saily","Nomad") and key=="std" and duration!="7 days":
                notes.append(f"{comp}: closest = {data}/{duration}")
        row.append("; ".join(notes))
        ws.append(row)
        for c in range(1, 11):
            cell = ws.cell(row=i, column=c)
            cell.border = BORDER
            cell.font = DEST_FONT if c==1 else BODY_FONT
            cell.alignment = LEFT
        prices = []
        for col_idx in (3,5,7,9):
            v = ws.cell(row=i, column=col_idx).value
            if isinstance(v,(int,float)):
                prices.append((col_idx, v))
                ws.cell(row=i, column=col_idx).number_format = '"$"#,##0.00'
        if prices:
            mn = min(p for _,p in prices); mx = max(p for _,p in prices)
            for col_idx, v in prices:
                if v==mn: ws.cell(row=i, column=col_idx).fill = GREEN_FILL
                elif v==mx: ws.cell(row=i, column=col_idx).fill = RED_FILL
    widths = {"A":18,"B":22,"C":14,"D":22,"E":14,"F":22,"G":14,"H":22,"I":14,"J":35}
    for col, w in widths.items(): ws.column_dimensions[col].width = w
    ws.freeze_panes = "B2"
    ws.row_dimensions[1].height = 32

ws1 = wb.active; ws1.title = "Standard 5GB 7d"; build_pricing(ws1, "std")
ws2 = wb.create_sheet("Heavy 10GB 30d"); build_pricing(ws2, "heavy")

ws3 = wb.create_sheet("Source URLs")
ws3.append(["Destination"]+COMPETITORS); style_header(ws3, 1, 5)
for i, dest in enumerate(DESTINATIONS, start=2):
    ws3.append([dest]+[url_for(c, dest) for c in COMPETITORS])
    for c in range(1, 6):
        cell = ws3.cell(row=i, column=c)
        cell.border = BORDER; cell.font = DEST_FONT if c==1 else BODY_FONT
        cell.alignment = LEFT
for col, w in [("A",18),("B",48),("C",48),("D",48),("E",48)]:
    ws3.column_dimensions[col].width = w
ws3.freeze_panes = "B2"

ws4 = wb.create_sheet("Summary")
sh = ["Destination","Tier","Std Avg ($)","Std Cheapest","Std Most Expensive","Std Spread ($)","Heavy Avg ($)","Heavy Cheapest","Heavy Most Expensive","Heavy Spread ($)","Peanut Target Std (-10%)","Peanut Target Heavy (-10%)"]
ws4.append(sh); style_header(ws4, 1, len(sh))
for i, dest in enumerate(DESTINATIONS, start=2):
    sp = {c: DATA[dest][c]["std"][3] for c in COMPETITORS}
    hp = {c: DATA[dest][c]["heavy"][3] for c in COMPETITORS}
    sa = sum(sp.values())/len(sp); ha = sum(hp.values())/len(hp)
    ws4.append([dest, TIERS[dest], sa, min(sp,key=sp.get), max(sp,key=sp.get), max(sp.values())-min(sp.values()), ha, min(hp,key=hp.get), max(hp,key=hp.get), max(hp.values())-min(hp.values()), sa*0.9, ha*0.9])
for i in range(2, len(DESTINATIONS)+2):
    for c in range(1, len(sh)+1):
        cell = ws4.cell(row=i, column=c); cell.border = BORDER
        cell.font = DEST_FONT if c==1 else BODY_FONT
        cell.alignment = CENTER
        if c in (3,6,7,10,11,12): cell.number_format = '"$"#,##0.00'
        if c in (4,8): cell.fill = GREEN_FILL
        if c in (5,9): cell.fill = RED_FILL
for col, w in [("A",16),("B",8),("C",14),("D",14),("E",18),("F",14),("G",14),("H",14),("I",18),("J",14),("K",22),("L",22)]:
    ws4.column_dimensions[col].width = w
ws4.freeze_panes = "B2"

ws5 = wb.create_sheet("Methodology")
notes = [
    ("Peanut eSIM Competitive Pricing Benchmark", True, 16),
    ("Generated: 2026-05-21 (Dubai time)", False, 10),
    ("", False, 10),
    ("Methodology", True, 12),
    ("- Plans scraped from competitor sites (Airalo, Holafly, Nomad, Saily) during May 2026.", False, 10),
    ("- All prices in USD as displayed by each site (US-default currency view).", False, 10),
    ("- Standard profile target: 5GB / 7 days. Closest match recorded for each competitor.", False, 10),
    ("- Heavy profile target: 10GB / 30 days. Closest match recorded.", False, 10),
    ("", False, 10),
    ("Key caveats", True, 12),
    ("- Holafly only sells UNLIMITED day-based plans (no GB tiers). Both profiles use Holafly's 7-day ($27.30) and 30-day ($74.90) unlimited plans. Customers get unlimited data — not a like-for-like GB price comparison.", False, 10),
    ("- Saily and Nomad do not offer 5GB / 7-day plans. Their closest match is 5GB / 30-day (most pricing is by data tier, not duration). Pricing in the Standard sheet reflects 30-day duration for these two competitors.", False, 10),
    ("- Airalo offers true 5GB/7d plans; Airalo data in the Standard sheet is the exact spec match.", False, 10),
    ("- All 22 destinations: every competitor offers service. No N/A cells in the matrix.", False, 10),
    ("", False, 10),
    ("Color coding", True, 12),
    ("- Green = cheapest competitor for that row. Red = most expensive.", False, 10),
    ("- Highlighted at price-cell level on Standard / Heavy sheets and on Cheapest / Most Expensive columns of Summary sheet.", False, 10),
]
for i, (text, bold, size) in enumerate(notes, start=1):
    ws5.cell(row=i, column=1, value=text)
    ws5.cell(row=i, column=1).font = Font(name="Arial", bold=bold, size=size)
ws5.column_dimensions["A"].width = 130

out = "/Users/a44/Downloads/peanut esim /Competitor_Pricing_Benchmark.xlsx"
wb.save(out)
print(f"Saved {out}")
