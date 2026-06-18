#!/usr/bin/env python3
"""
Build the eSIM Pricing Explorer data layer from raw supplier catalogs.

SINGLE SOURCE OF TRUTH (read-only inputs):
  - eSIM-Access-Price.csv           (eSIM Access, USD pass-through, ISO2 in `Code`)
  - Samurai-Wifi-Wholesale-CNY.xlsx (Samurai, CNY settlement -> USD via FX_CNY_PER_USD)

OUTPUT:
  - Regenerates the inline `const SEARCH_INDEX`, `const DATA`, `const FLAGS`
    lines in pricing-dashboard.html (a .bak is written first).
  - Also drops tools/pricing-data.json for inspection / future fetch use.

Cross-supplier join key is ISO2 -> one canonical display name, so EA and
Samurai SKUs for the same country share an identical `c` string and line up.

Re-run after exporting a fresh eSIM Access CSV. Never hand-edit the data in
the HTML again.
"""
import csv, json, re, sys, shutil, collections
from pathlib import Path
import openpyxl
import pycountry

# ----------------------------------------------------------------------------
# CONFIG
# ----------------------------------------------------------------------------
FX_CNY_PER_USD = 7.149  # PLACEHOLDER — Samurai settles in CNY. Correct to the
                        # contractual rate; every Samurai USD price scales by this.

SRC = Path("/Users/a44/Documents/dojo/dojo/Project-Master vault /Peanut eSIM/04-Pricing-Supply")
EA_CSV   = SRC / "eSIM-Access-Price.csv"
SAM_XLSX = SRC / "Samurai-Wifi-Wholesale-CNY.xlsx"
TELNA_CSV   = SRC / "Telna-Connect-Flex-Products.csv"   # Telna, USD direct, ISO3 lists
ESIMGO_XLSX = SRC / "eSIM-Go-Rate-Sheet-Standard-CONFIDENTIAL.xlsx"  # eSIM Go, USD
ZETEXA_XLSX = SRC / "Zetexa-Pricing-CONFIDENTIAL.xlsx"  # Zetexa, USD direct, NDA'd

REPO = Path.home() / "code/peanut-esim-business-overview"
HTML = REPO / "tools/pricing-dashboard.source.html"   # unencrypted working master (gitignored).
JSON_OUT = REPO / "tools/pricing-data.json"            # the deployed pricing-dashboard.html is the
                                                       # staticrypt-ENCRYPTED build of this source — see BUILD.md.

# ----------------------------------------------------------------------------
# CANONICAL COUNTRY NAMES  (ISO2 -> display)   built from EA's clean Code->Region
# with overrides to match the dashboard's Emirates canon (FLIGHTS/SEARCH_INDEX).
# ----------------------------------------------------------------------------
ISO_NAME_OVERRIDE = {
    "HK": "Hong Kong", "MO": "Macao", "CI": "Ivory Coast", "CN": "China mainland",
    "KR": "South Korea", "TW": "Taiwan", "US": "United States", "GB": "United Kingdom",
    "AE": "United Arab Emirates", "RU": "Russia", "LA": "Laos", "MP": "Saipan",
    "GU": "Guam",
}
NAME_BY_ISO = {}  # filled from EA, then overridden

def flag_emoji(iso2):
    if not iso2 or len(iso2) != 2 or not iso2.isalpha():
        return ""
    return "".join(chr(0x1F1E6 + ord(c) - ord("A")) for c in iso2.upper())

def canon_name(iso2):
    if iso2 in ISO_NAME_OVERRIDE:
        return ISO_NAME_OVERRIDE[iso2]
    if iso2 in NAME_BY_ISO:
        return NAME_BY_ISO[iso2]
    c = pycountry.countries.get(alpha_2=iso2)
    return (getattr(c, "common_name", None) or c.name) if c else iso2

# ----------------------------------------------------------------------------
# SAMURAI commodity-name -> classification.
# Each distinct leading token maps to either:
#   ("S", "ISO2")          single country
#   ("B", ["ISO2", ...])   multi-country bundle (covers)
#   ("X", None)            skip (junk / service charge)
# Region bundles list their Emirates-relevant members (best-effort, not exhaustive).
# ----------------------------------------------------------------------------
GCC   = ["AE","SA","KW","QA","BH","OM"]
SEA   = ["SG","MY","TH","ID","VN","PH","KH","LA","MM","BN"]
EURO  = ["GB","IE","FR","DE","IT","ES","PT","NL","BE","AT","CH","DK","SE","NO","FI",
         "PL","CZ","HU","GR","RO","HR","SI","SK","BG","LU","MT","CY","EE","LV","LT","IS"]
NAM   = ["US","CA","MX"]
AFR   = ["ZA","KE","TZ","EG","MA","NG","GH","UG","ZM","SN","DZ","TN","ET"]
ASIA  = ["JP","KR","CN","HK","TW","SG","MY","TH","ID","VN","PH","IN","KH","LA","MO"]
LATAM = ["BR","AR","CL","CO","PE","UY","BO","EC","PY"]
BALK  = ["RS","ME","BA","MK","AL"]
MENA  = GCC + ["EG","JO","LB","IL","TR","MA","TN","DZ","IQ"]

SAM_TOKENS = {
    # --- singles ---
    "japan":"S:JP", "china mainland":"S:CN", "china":"S:CN", "vietnam":"S:VN",
    "uae":"S:AE", "hong kong":"S:HK", "usa":"S:US", "korea":"S:KR",
    "south korea":"S:KR", "pakistan":"S:PK", "saudi arabia":"S:SA", "egypt":"S:EG",
    "mexico":"S:MX", "new zealand":"S:NZ", "brazil":"S:BR", "south africa":"S:ZA",
    "kenya":"S:KE", "singapore":"S:SG", "india":"S:IN", "maldives":"S:MV",
    "tanzania":"S:TZ", "taiwan":"S:TW", "turkey":"S:TR", "canada":"S:CA",
    "thailand":"S:TH",
    # --- multi-country bundles with explicit members in the name ---
    "china hong kong macao":"B:CN,HK,MO", "hong kong macao":"B:HK,MO",
    "us canada mexico":"B:US,CA,MX", "australia new zealand":"B:AU,NZ",
    "singapore malaysia indonesia":"B:SG,MY,ID", "uk ireland":"B:GB,IE",
    "denmark sweden":"B:DK,SE", "jordan kuwait oman":"B:JO,KW,OM",
    "brazil chile":"B:BR,CL",
    "kazakhstan uzbekistan kyrgyzstan pakistan":"B:KZ,UZ,KG,PK",
    "guam saipan":"B:GU,MP",
    # --- region bundles (token = name AFTER trailing counts are stripped) ---
    "europe":"B:"+",".join(EURO), "new europe":"B:"+",".join(EURO),
    "orange holiday europe":"B:"+",".join(EURO),
    "southeast asia":"B:"+",".join(SEA),
    "north america":"B:"+",".join(NAM), "africa":"B:"+",".join(AFR),
    "me & africa":"B:"+",".join(MENA+AFR),
    "asia":"B:"+",".join(ASIA), "asialink":"B:"+",".join(ASIA),
    "asian multiple regions":"B:"+",".join(ASIA),
    "gulf arab":"B:"+",".join(GCC), "eu+mena":"B:"+",".join(EURO+MENA),
    "western balkans":"B:"+",".join(BALK),
    "south america":"B:"+",".join(LATAM),
    # global bundles: no specific country filter
    "global":"B:", "global156":"B:", "global88":"B:", "orange world global":"B:",
    # --- junk ---
    "service charge":"X:",
}

# decoration strippers applied before tokenizing a Samurai name
DECO = [
    (re.compile(r"\[[^\]]*\]"), ""),                       # [5G]
    (re.compile(r"\((?:plus|rnr|multi|skt|iij|softbank|kddi|ooredoo|ctm)\)", re.I), ""),
    (re.compile(r"^\s*\(?(?:plus|rnr|co)\)?\s+", re.I), ""),  # leading (plus)/(RNR)/CO
    (re.compile(r"\(china\)", re.I), ""),
    (re.compile(r"\s+viettel.*$", re.I), ""),              # Vietnam Viettel Tour80
    (re.compile(r"\s+(dtac|true)\s+\d+.*$", re.I), ""),    # Thailand Dtac 219
    (re.compile(r"^\s*esim carrier of \d+ days\s*\+\s*", re.I), ""),  # "...+Korea"
    (re.compile(r"^5g\]\s*", re.I), ""),                   # stray "5G] Japan"
]

def sam_token(name):
    s = name
    for rx, repl in DECO:
        s = rx.sub(repl, s)
    s = s.strip()
    # geo token = text before first '-' or ',' or digit-group; collapse spaces
    tok = re.split(r"[-,]", s)[0]
    tok = re.sub(r"\s+\d.*$", "", tok)      # drop trailing numbers (counts)
    tok = re.sub(r"\s+", " ", tok).strip().lower()
    return tok

def classify_sam(name):
    tok = sam_token(name)
    if tok in SAM_TOKENS:
        kind, payload = SAM_TOKENS[tok].split(":", 1)
        if kind == "S":
            return ("S", payload)
        if kind == "B":
            return ("B", [c for c in payload.split(",") if c])
        return ("X", None)
    # fallback: try a single-country pycountry match on the token
    try:
        hit = pycountry.countries.lookup(tok)
        return ("S", hit.alpha_2)
    except LookupError:
        return ("?", tok)

# ----------------------------------------------------------------------------
# helpers
# ----------------------------------------------------------------------------
def money(s):
    return round(float(re.sub(r"[^\d.]", "", str(s)) or 0), 4)

def data_label(gb):
    if not gb or gb <= 0:
        return "Unlimited"
    if gb < 1:
        return f"{int(round(gb*1000))}MB"
    return f"{int(gb)}GB" if float(gb).is_integer() else f"{gb}GB"

def parse_hs_gb(hs):
    """Samurai 'High Speed Data' -> GB float. '10.00GB' / '1.00GB/Day' / '0.00MB/Day'."""
    m = re.search(r"([\d.]+)\s*(GB|MB)", str(hs), re.I)
    if not m:
        return 0.0
    v = float(m.group(1))
    # Samurai expresses daily caps in MB (1024 MB == 1 GB); use the binary divisor
    # so "1024.00MB" -> 1.0 GB cleanly instead of 1.024.
    return v if m.group(2).upper() == "GB" else v/1024.0

def iso3_to_iso2(code):
    """'NGA' -> 'NG'. Returns None for unknown / non-ISO codes (e.g. 'KSV','ANT')."""
    code = (code or "").strip().upper()
    if len(code) == 2:
        return code
    try:
        c = pycountry.countries.get(alpha_3=code) or pycountry.countries.lookup(code)
        return c.alpha_2
    except (LookupError, AttributeError):
        return None

# ----------------------------------------------------------------------------
# 1) eSIM Access
# ----------------------------------------------------------------------------
def build_ea():
    rows = list(csv.DictReader(open(EA_CSV, encoding="utf-8-sig")))
    # learn ISO2 -> region name from singles (clean source of canonical names)
    for x in rows:
        if x["Type"] == "Single" and len(x["Code"]) == 2:
            NAME_BY_ISO.setdefault(x["Code"], x["Region"])
    singles, bundles = [], []
    multi = collections.defaultdict(list)
    for x in rows:
        if x["Type"] == "Single":
            iso = x["Code"]
            gb = money(x["GBs"]) if x["GBs"] not in ("", None) else 0
            daily = x["Data Type"] == "Daily Unlimited"
            singles.append({
                "s":"EA", "c":canon_name(iso), "d":data_label(gb), "gv":gb,
                "sp":"Daily refill" if daily else "Fixed data",
                "pd":int(float(x["Pre-Install Days"] or 0)) if str(x["Pre-Install Days"]).strip().isdigit() else 180,
                "v":int(float(x["Validity(Days)"] or 0)), "p":money(x["Price(USD)"]),
                "n":x["Name"], "t":"S",
            })
        else:  # Multi-Area -> bundle
            multi[x["Region"]].append(x)
    for region, xs in multi.items():
        covers = []
        for code in (xs[0]["Coverage"] or "").split(","):
            code = code.strip()
            if len(code) == 2:
                nm = canon_name(code)
                if nm not in covers:
                    covers.append(nm)
        skus = [{
            "s":"EA", "b":region, "d":data_label(money(x["GBs"]) if x["GBs"] else 0),
            "v":int(float(x["Validity(Days)"] or 0)), "p":money(x["Price(USD)"]), "n":x["Name"],
        } for x in xs]
        bundles.append({"name":region, "supplier":"EA", "covers":covers, "skus":skus})
    return singles, bundles

# ----------------------------------------------------------------------------
# 2) Samurai
# ----------------------------------------------------------------------------
def build_sam():
    wb = openpyxl.load_workbook(SAM_XLSX, read_only=True, data_only=True)
    singles, bundles = [], []
    sam_bundles = {}      # name -> bundle dict
    unresolved = collections.Counter()
    for sn in wb.sheetnames:
        ws = wb[sn]
        header = None
        for row in ws.iter_rows(min_row=1, values_only=True):
            if row and row[0] == "No.":
                header = [str(h).strip() if h else "" for h in row]
                break
        idx = {h:i for i,h in enumerate(header)}
        def col(r, name):
            i = idx.get(name)
            return r[i] if i is not None and i < len(r) else None
        for r in ws.iter_rows(min_row=1, values_only=True):
            if not r or r[0] in (None, "No.", "Purchase goods and price information"):
                continue
            name = str(col(r, "Commodity Name") or "").strip()
            settle = col(r, "Settlement price")
            days = col(r, "Days")
            if not name or settle in (None, "") or days in (None, ""):
                continue
            try:
                days = int(float(days)); p = round(float(settle)/FX_CNY_PER_USD, 3)
            except (TypeError, ValueError):
                continue
            gb = parse_hs_gb(col(r, "High Speed Data"))
            thr = str(col(r, "Throttled Speed (kbps)") or "").strip()
            sp = f"{thr}kbps" if thr and thr not in ("0","") else ""
            pdm = re.search(r"of\s+(\d+)\s*days", name, re.I)
            pd = int(pdm.group(1)) if pdm else 90
            kind, payload = classify_sam(name)
            if kind == "X":
                continue
            if kind == "S":
                singles.append({
                    "s":"Sam", "c":canon_name(payload), "d":data_label(gb), "gv":gb,
                    "sp":sp, "pd":pd, "v":days, "p":p, "n":name, "t":"P",
                })
            elif kind == "B":
                b = sam_bundles.get(name)
                if not b:
                    covers = [canon_name(c) for c in payload]
                    b = {"name":name, "supplier":"Sam", "covers":covers, "skus":[]}
                    sam_bundles[name] = b
                b["skus"].append({"s":"Sam","b":name,"d":data_label(gb),"v":days,"p":p,"n":name})
            else:
                unresolved[payload] += 1
    bundles = list(sam_bundles.values())
    return singles, bundles, unresolved

# ----------------------------------------------------------------------------
# 3) Telna  (Connect Flex catalog — USD direct, ISO3 country lists, MB allowance)
# ----------------------------------------------------------------------------
def build_telna():
    rows = list(csv.DictReader(open(TELNA_CSV, encoding="utf-8-sig")))
    singles, bundles = [], []
    for x in rows:
        name = (x.get("Product Name") or "").strip()
        if not name:
            continue
        p = money(x.get("Cost (USD)"))
        gb = round(money(x.get("Data Allowance (MB)")) / 1024.0, 4)  # MB -> GB
        dm = re.search(r"(\d+)", str(x.get("Duration") or ""))
        days = int(dm.group(1)) if dm else 0
        iso2s = [i for i in (iso3_to_iso2(c) for c in (x.get("Countries") or "").split(",")) if i]
        if len(iso2s) == 1:
            singles.append({
                "s":"Telna", "c":canon_name(iso2s[0]), "d":data_label(gb), "gv":gb,
                "sp":"Fixed data", "pd":0, "v":days, "p":p, "n":name, "t":"S",
            })
        elif len(iso2s) > 1:
            covers, seen = [], set()
            for i in iso2s:
                nm = canon_name(i)
                if nm not in seen:
                    seen.add(nm); covers.append(nm)
            bundles.append({"name":name, "supplier":"Telna", "covers":covers,
                "skus":[{"s":"Telna","b":name,"d":data_label(gb),"v":days,"p":p,"n":name}]})
    return singles, bundles

# ----------------------------------------------------------------------------
# 4) eSIM Go  (Standard - Fixed sheet: per-country fixed bundles, USD)
#    Each country row carries 6 GB/validity tiers across the price columns.
#    Region rows (ISOCode with ';') -> multi-country bundles.
# ----------------------------------------------------------------------------
def build_esimgo():
    wb = openpyxl.load_workbook(ESIMGO_XLSX, read_only=True, data_only=True)
    singles, bundles = [], []
    ws = wb["Standard - Fixed"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return singles, bundles
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    # price columns look like "1GB/7days (USD)" -> (col_index, gb, days)
    tiers = []
    for i, h in enumerate(header):
        m = re.search(r"(\d+)\s*GB\s*/\s*(\d+)\s*days", h, re.I)
        if m:
            tiers.append((i, int(m.group(1)), int(m.group(2))))
    try:
        iso_col = header.index("ISOCode")
        ctry_col = header.index("Country")
    except ValueError:
        return singles, bundles
    for r in rows[1:]:
        if not r or len(r) <= iso_col:
            continue
        iso_raw = str(r[iso_col] or "").strip()
        country = str(r[ctry_col] or "").strip()
        if not iso_raw or not country:
            continue
        iso_list = [c.strip() for c in iso_raw.split(";") if c.strip()]
        is_bundle = len(iso_list) > 1
        if is_bundle:
            covers, seen = [], set()
            for c in iso_list:
                i2 = iso3_to_iso2(c)
                if not i2:
                    continue
                nm = canon_name(i2)
                if nm not in seen:
                    seen.add(nm); covers.append(nm)
            skus = []
            for ci, gb, days in tiers:
                if ci < len(r) and r[ci] not in (None, "", "NA"):
                    skus.append({"s":"EGo","b":country,"d":data_label(gb),"v":days,
                                 "p":money(r[ci]),"n":f"{country} {gb}GB/{days}d"})
            if skus:
                bundles.append({"name":country, "supplier":"EGo", "covers":covers, "skus":skus})
        else:
            i2 = iso3_to_iso2(iso_list[0]) if iso_list else None
            if not i2:
                continue
            for ci, gb, days in tiers:
                if ci < len(r) and r[ci] not in (None, "", "NA"):
                    singles.append({
                        "s":"EGo", "c":canon_name(i2), "d":data_label(gb), "gv":float(gb),
                        "sp":"Fixed data", "pd":0, "v":days, "p":money(r[ci]),
                        "n":f"{country} {gb}GB/{days}d", "t":"S",
                    })
    return singles, bundles

# ----------------------------------------------------------------------------
# 5) Zetexa  (single "Pricing" sheet, region-grouped: 6 fixed tiers
#    (1GB/1d … 50GB/30d) + 6 unlimited tiers (3d … 20d) per country across the
#    columns; USD direct; "—" = tier not offered; "N Countries (...)" rows are
#    regional bundles. CONFIDENTIAL (NDA) — never commit the raw sheet/JSON.
# ----------------------------------------------------------------------------
ZETEXA_NAME_OVERRIDE = {
    "south korea":"KR", "russian federation":"RU", "hong kong":"HK",
    "congo dem. rep":"CD", "congo republic":"CG", "czech republic":"CZ",
    "vietnam":"VN", "vitenam":"VN", "turkey":"TR", "cape verde":"CV",
    "french guiana":"GF", "french polynesia":"PF", "faroe islands":"FO",
    "curacao":"CW", "moldova":"MD", "taiwan":"TW", "macao":"MO", "macau":"MO",
    "laos":"LA", "ivory coast":"CI", "brunei":"BN", "syria":"SY", "iran":"IR",
    "united states":"US", "united kingdom":"GB", "bolivia":"BO",
    "bosnia and herzegovina":"BA", "macao china":"MO", "macedonia":"MK",
    "palestinian territory":"PS", "reunion":"RE",
    "saint vincent and grenadines":"VC",
    # "Netherlands Antilles" intentionally NOT mapped — defunct ISO grouping
    # (dissolved 2010); resolves to nothing → dropped (1 obscure SKU set).
}

def ztx_name_iso2(name):
    n = re.sub(r"\s+", " ", (name or "").strip())
    key = n.lower()
    if key in ZETEXA_NAME_OVERRIDE:
        return ZETEXA_NAME_OVERRIDE[key]
    try:
        return pycountry.countries.lookup(n).alpha_2
    except LookupError:
        return None

def parse_ztx_tier(h):
    """'10GB/10 Days' -> (10.0, 10, False); 'Unlim 7 Days' -> (0.0, 7, True)."""
    h = str(h or "").strip()
    dm = re.search(r"(\d+)\s*Days?", h, re.I)
    if not dm:
        return None
    unlim = bool(re.search(r"unlim", h, re.I))
    gbm = re.search(r"(\d+)\s*GB", h, re.I)
    gb = 0.0 if unlim else (float(gbm.group(1)) if gbm else 0.0)
    return (gb, int(dm.group(1)), unlim)

def build_zetexa():
    if not ZETEXA_XLSX.exists():
        return [], [], collections.Counter()
    wb = openpyxl.load_workbook(ZETEXA_XLSX, read_only=True, data_only=True)
    ws = wb["Pricing"]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    tiers = []                                   # (col_index, gb, days, unlim)
    for i in range(2, len(header)):
        t = parse_ztx_tier(header[i])
        if t:
            tiers.append((i, *t))
    singles, bundles = [], []
    unresolved = collections.Counter()
    bundle_rx = re.compile(r"^\s*\d+\s+countr(?:y|ies)\s*\(", re.I)

    def price_of(r, ci):
        if ci >= len(r):
            return None
        v = r[ci]
        if v in (None, "", "—", "-", "–"):
            return None
        try:
            return round(float(v), 4)
        except (TypeError, ValueError):
            return None

    for r in rows[1:]:
        if not r:
            continue
        region  = str(r[0] or "").strip()
        country = str(r[1] or "").strip() if len(r) > 1 else ""
        if not country or country.startswith("▸"):
            continue                              # region-header / spacer rows
        if bundle_rx.match(country):              # regional bundle
            inside = re.search(r"\(([^)]*)\)", country)
            members = [m.strip() for m in (inside.group(1).split(",") if inside else []) if m.strip()]
            covers, seen = [], set()
            for m in members:
                iso = ztx_name_iso2(m)
                if not iso:
                    unresolved[m] += 1; continue
                nm = canon_name(iso)
                if nm not in seen:
                    seen.add(nm); covers.append(nm)
            bname = f"Zetexa {region} ({len(covers)} countries)"
            skus = []
            for ci, gb, days, unlim in tiers:
                p = price_of(r, ci)
                if p is not None:
                    skus.append({"s":"Ztx","b":bname,"d":data_label(gb),"v":days,"p":p,
                                 "n":f"{bname} {data_label(gb)}/{days}d"})
            if skus and covers:
                bundles.append({"name":bname,"supplier":"Ztx","covers":covers,"skus":skus})
            continue
        iso = ztx_name_iso2(country)              # single country
        if not iso:
            unresolved[country] += 1; continue
        for ci, gb, days, unlim in tiers:
            p = price_of(r, ci)
            if p is not None:
                singles.append({
                    "s":"Ztx", "c":canon_name(iso), "d":data_label(gb), "gv":gb,
                    "sp":"Unlimited" if unlim else "Fixed data", "pd":0, "v":days,
                    "p":p, "n":f"{country} {data_label(gb)}/{days}d", "t":"S",
                })
    return singles, bundles, unresolved

# ----------------------------------------------------------------------------
# main
# ----------------------------------------------------------------------------
def main():
    ea_s, ea_b = build_ea()
    sam_s, sam_b, unresolved = build_sam()
    tel_s, tel_b = build_telna()
    ego_s, ego_b = build_esimgo()
    ztx_s, ztx_b, ztx_unresolved = build_zetexa()
    singles = sam_s + ea_s + ego_s + tel_s + ztx_s   # Samurai-first ordering preserved
    bundles = sam_b + ea_b + ego_b + tel_b + ztx_b

    data = {"singles": singles, "bundles": bundles}

    # FLAGS for every canonical single country
    countries = sorted({d["c"] for d in singles})
    iso_by_name = {}
    for c in pycountry.countries:                       # full reverse map (covers Zetexa long-tail)
        iso_by_name.setdefault(canon_name(c.alpha_2), c.alpha_2)
    for iso in set(NAME_BY_ISO) | set(ISO_NAME_OVERRIDE) | {"TW","MP","GU"}:
        iso_by_name[canon_name(iso)] = iso             # overrides win
    flags = {c: flag_emoji(iso_by_name.get(c, "")) for c in countries if flag_emoji(iso_by_name.get(c, ""))}

    # SEARCH_INDEX entry per country
    search = [{"key":c.lower(), "label":c, "country":c} for c in countries]

    # ---- report ----
    ea_cty  = len({d["c"] for d in ea_s})
    sam_cty = len({d["c"] for d in sam_s})
    both = {d["c"] for d in ea_s} & {d["c"] for d in sam_s}
    print("="*64)
    print("GENERATED PRICING DATA")
    print(f"  EA   singles: {len(ea_s):>6}   countries: {ea_cty}")
    print(f"  Sam  singles: {len(sam_s):>6}   countries: {sam_cty}")
    print(f"  EGo  singles: {len(ego_s):>6}   countries: {len({d['c'] for d in ego_s})}")
    print(f"  Telna singles:{len(tel_s):>6}   countries: {len({d['c'] for d in tel_s})}")
    print(f"  Ztx  singles: {len(ztx_s):>6}   countries: {len({d['c'] for d in ztx_s})}")
    print(f"  bundles -> EA:{len(ea_b)}  Sam:{len(sam_b)}  EGo:{len(ego_b)}  Telna:{len(tel_b)}  Ztx:{len(ztx_b)}")
    if ztx_unresolved:
        print("  !! UNRESOLVED Zetexa country names (add to ZETEXA_NAME_OVERRIDE):")
        for nm, n in ztx_unresolved.most_common():
            print(f"       {n:>4}  {nm!r}")
    print(f"  countries with BOTH EA+Sam (singles): {len(both)}")
    print(f"  total single records: {len(singles)}   total bundles: {len(bundles)}")
    print(f"  FX: settlement CNY / {FX_CNY_PER_USD} = USD")
    if unresolved:
        print("  !! UNRESOLVED Samurai tokens (add to SAM_TOKENS):")
        for tok, n in unresolved.most_common():
            print(f"       {n:>4}  {tok!r}")
    else:
        print("  all Samurai commodities classified ✓")
    print("="*64)

    JSON_OUT.write_text(json.dumps({**data, "flags":flags, "search":search,
        "meta":{"fx_cny_per_usd":FX_CNY_PER_USD,
                "source":["eSIM-Access-Price.csv","Samurai-Wifi-Wholesale-CNY.xlsx",
                          "eSIM-Go-Rate-Sheet-Standard-CONFIDENTIAL.xlsx",
                          "Telna-Connect-Flex-Products.csv",
                          "Zetexa-Pricing-CONFIDENTIAL.xlsx"]}},
        ensure_ascii=False))
    print(f"wrote {JSON_OUT}  ({JSON_OUT.stat().st_size/1e6:.2f} MB)")

    if "--patch" in sys.argv:
        patch_html(data, flags, search)

def patch_html(data, flags, search):
    txt = HTML.read_text().splitlines(keepends=True)
    shutil.copy(HTML, HTML.with_suffix(".html.bak"))
    def js(name, obj):
        return f"const {name} = {json.dumps(obj, ensure_ascii=False, separators=(',',':'))};\n"
    out, i, n = [], 0, len(txt)
    replaced = collections.Counter()
    while i < n:
        line = txt[i]
        # Only DATA + FLAGS are regenerated. SEARCH_INDEX is left untouched to
        # preserve its hand-curated airport-code / city-alias search entries.
        m = re.match(r"\s*const (DATA|FLAGS)\s*=", line)
        if m:
            name = m.group(1)
            # consume until the statement terminates (line ending with '};' or '];')
            j = i
            while j < n and not re.search(r"[\]}]\s*;\s*$", txt[j]):
                j += 1
            payload = {"SEARCH_INDEX":search, "DATA":data, "FLAGS":flags}[name]
            out.append(js(name, payload))
            replaced[name] += 1
            i = j + 1
            continue
        out.append(line)
        i += 1
    HTML.write_text("".join(out))
    print("patched HTML — replaced:", dict(replaced), " (backup: pricing-dashboard.html.bak)")

if __name__ == "__main__":
    main()
