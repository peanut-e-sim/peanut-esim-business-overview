#!/usr/bin/env python3
"""Validate generated pricing-data.json against the raw source catalogs."""
import csv, json, re, collections
from pathlib import Path
import openpyxl

SRC = Path("/Users/a44/Documents/dojo/dojo/Project-Master vault /Peanut eSIM/04-Pricing-Supply")
EA_CSV = SRC / "eSIM-Access-Price.csv"
SAM_XLSX = SRC / "Samurai-Wifi-Wholesale-CNY.xlsx"
DATA = json.loads((Path.home()/"code/peanut-esim-business-overview/tools/pricing-data.json").read_text())
FX = DATA["meta"]["fx_cny_per_usd"]

singles, bundles = DATA["singles"], DATA["bundles"]
ea_s  = [d for d in singles if d["s"]=="EA"]
sam_s = [d for d in singles if d["s"]=="Sam"]
ea_b  = [b for b in bundles if b["supplier"]=="EA"]
sam_b = [b for b in bundles if b["supplier"]=="Sam"]
ea_b_skus  = sum(len(b["skus"]) for b in ea_b)
sam_b_skus = sum(len(b["skus"]) for b in sam_b)

# ---- source counts ----
ea_rows = list(csv.DictReader(open(EA_CSV, encoding="utf-8-sig")))
ea_single_src = sum(1 for x in ea_rows if x["Type"]=="Single")
ea_multi_src  = sum(1 for x in ea_rows if x["Type"]=="Multi-Area")

wb = openpyxl.load_workbook(SAM_XLSX, read_only=True, data_only=True)
sam_valid, sam_service = 0, 0
for sn in wb.sheetnames:
    ws = wb[sn]
    hdr=None
    for row in ws.iter_rows(values_only=True):
        if row and row[0]=="No.": hdr=[str(h).strip() if h else "" for h in row]; break
    idx={h:i for i,h in enumerate(hdr)}
    for r in ws.iter_rows(values_only=True):
        if not r or r[0] in (None,"No.","Purchase goods and price information"): continue
        nm=str(r[idx["Commodity Name"]] or "").strip()
        st=r[idx["Settlement price"]]; dy=r[idx["Days"]]
        if not nm or st in (None,"") or dy in (None,""): continue
        try: int(float(dy)); float(st)
        except: continue
        if nm.lower().startswith("service charge"): sam_service+=1; continue
        sam_valid+=1

def check(label, ok, detail=""):
    print(f"  [{'PASS' if ok else 'FAIL'}] {label}{'  '+detail if detail else ''}")
    return ok

print("="*70); print("VALIDATION"); print("="*70)
allok=True

print("\nCheck 1 — eSIM Access count parity")
allok &= check(f"EA singles {len(ea_s)} == source Single rows {ea_single_src}", len(ea_s)==ea_single_src)
allok &= check(f"EA bundle SKUs {ea_b_skus} == source Multi-Area rows {ea_multi_src}", ea_b_skus==ea_multi_src)
allok &= check(f"EA countries {len({d['c'] for d in ea_s})} == 186", len({d['c'] for d in ea_s})==186)

print("\nCheck 2 — Samurai count parity (every valid row -> single or bundle sku)")
gen_sam = len(sam_s)+sam_b_skus
allok &= check(f"Sam generated {gen_sam} == source valid rows {sam_valid} (service-charge skipped: {sam_service})", gen_sam==sam_valid)

print("\nCheck 3 — Cote d'Ivoire / Ivory Coast fix")
names={d["c"] for d in singles}
allok &= check("'Ivory Coast' present", "Ivory Coast" in names)
allok &= check("'Cote d'Ivoire' NOT present (normalized away)", "Cote d'Ivoire" not in names)
ci=[d for d in ea_s if d["c"]=="Ivory Coast"]
allok &= check(f"Ivory Coast has EA SKUs ({len(ci)})", len(ci)>0)

print("\nCheck 4 — spot price accuracy (vs source)")
def find(s,c,gv,v,d=None):
    for x in singles:
        if x["s"]==s and x["c"]==c and abs(x["gv"]-gv)<1e-6 and x["v"]==v: return x
    return None
# EA China 1GB/day v1 = $0.65 ; EA Japan 10GB/day v1 = $5.00
allok &= check("EA China mainland 1GB v1 = 0.65", (find("EA","China mainland",1,1) or {}).get("p")==0.65, str((find("EA","China mainland",1,1) or {}).get("p")))
allok &= check("EA Japan 10GB v1 = 5.0", (find("EA","Japan",10,1) or {}).get("p")==5.0, str((find("EA","Japan",10,1) or {}).get("p")))
# Samurai USA 1GB: source settlement 4.2 CNY (day1) 5.5 CNY (day2) -> /FX
usa1=round(4.2/FX,3); usa2=round(5.5/FX,3)
u1=find("Sam","United States",1,1); u2=find("Sam","United States",1,2)
allok &= check(f"Sam USA 1GB v1 = {usa1} (4.2 CNY/FX)", bool(u1) and abs(u1["p"]-usa1)<0.01, str(u1 and u1["p"]))
allok &= check(f"Sam USA 1GB v2 = {usa2} (5.5 CNY/FX)", bool(u2) and abs(u2["p"]-usa2)<0.01, str(u2 and u2["p"]))

print("\nCheck 5 — schema integrity")
sk_ok=all(set(d)>= {"s","c","d","gv","sp","pd","v","p","n","t"} for d in singles[:5000])
b_ok=all(set(b)=={"name","supplier","covers","skus"} for b in bundles)
allok &= check("single record keys match dashboard schema", sk_ok)
allok &= check("bundle record keys match dashboard schema", b_ok)

print("\n"+"="*70)
print("OVERALL:", "PASS ✅" if allok else "FAIL ❌")
print("="*70)
